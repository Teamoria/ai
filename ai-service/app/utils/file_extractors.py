"""File text extraction helpers."""

from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
import re
import unicodedata
from urllib.parse import urlparse
from urllib.error import URLError
from urllib.request import Request, urlopen

from fastapi import HTTPException, status

from app.core.config import settings


TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".srt", ".vtt", ".log"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx"}
SPREADSHEET_EXTENSIONS = {".xlsx"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp"}
MEDIA_EXTENSIONS = {".mp3", ".mp4", ".mpeg", ".mpga", ".m4a", ".wav", ".webm", ".mov", ".avi", ".mkv"}


@dataclass(frozen=True)
class UploadSource:
    source_type: str
    text: str | None = None
    path: Path | None = None
    source: str = "content"


def resolve_upload_source(
    *,
    content: str | None = None,
    file_path: str | None = None,
    file_url: str | None = None,
    file_url_headers: dict[str, str] | None = None,
    file_url_api_key: str | None = None,
    file_url_bearer_token: str | None = None,
) -> UploadSource:
    if content:
        return UploadSource(source_type="text", text=clean_extracted_text(content), source="content")

    if file_path:
        path = Path(file_path)

        if path.exists() and path.is_file():
            return _source_from_path(path, source=file_path)

        if not file_url:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="file_path does not exist or is not a file.",
            )

    if file_url:
        return _source_from_url(
            file_url,
            headers=file_url_headers,
            api_key=file_url_api_key,
            bearer_token=file_url_bearer_token,
        )

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="Provide content, file_path, or file_url for processing.",
    )


def extract_text_from_source(
    *,
    content: str | None = None,
    file_path: str | None = None,
    file_url: str | None = None,
) -> str:
    """Extract text from raw content, a local path, or a URL.

    Laravel should prefer sending raw text when it already extracted it. For local
    demos this also supports plain text-like files. Binary document parsing can be
    added later without changing the API contract.
    """

    source = resolve_upload_source(content=content, file_path=file_path, file_url=file_url)

    if source.text is not None:
        return source.text

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"{source.source_type} uploads require a dedicated processor.",
    )


def _source_from_path(path: Path, *, source: str) -> UploadSource:
    suffix = path.suffix.lower()

    if suffix in MEDIA_EXTENSIONS:
        return UploadSource(source_type="media", path=path, source=source)

    if suffix in DOCUMENT_EXTENSIONS:
        return UploadSource(source_type=suffix.removeprefix("."), text=_read_document(path), source=source)

    if suffix in SPREADSHEET_EXTENSIONS:
        return UploadSource(source_type="xlsx", text=_read_spreadsheet(path), source=source)

    if suffix in IMAGE_EXTENSIONS:
        return UploadSource(source_type="image", text=_read_image(path), source=source)

    if suffix in TEXT_EXTENSIONS or suffix == "":
        return UploadSource(source_type="text", text=clean_extracted_text(path.read_text(encoding="utf-8", errors="ignore")), source=source)

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"Unsupported upload file type: {suffix}",
    )


def _source_from_url(
    file_url: str,
    *,
    headers: dict[str, str] | None = None,
    api_key: str | None = None,
    bearer_token: str | None = None,
) -> UploadSource:
    path = _download_url(
        file_url,
        headers=headers,
        api_key=api_key,
        bearer_token=bearer_token,
    )
    return _source_from_path(path, source=file_url)


def _download_url(
    file_url: str,
    *,
    headers: dict[str, str] | None = None,
    api_key: str | None = None,
    bearer_token: str | None = None,
) -> Path:
    suffix = Path(urlparse(file_url).path).suffix
    request = Request(file_url, headers=_download_headers(headers, api_key, bearer_token))

    try:
        with urlopen(request, timeout=20) as response:
            with NamedTemporaryFile(delete=False, suffix=suffix or ".upload") as temp_file:
                temp_file.write(response.read())
                return Path(temp_file.name)
    except (OSError, URLError) as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Unable to read file_url: {exc}",
        ) from exc


def _read_document(path: Path) -> str:
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        text = _read_pdf_with_pypdf(path)
        if _is_low_quality_text(text):
            fallback_text = _read_pdf_with_pymupdf(path)
            if fallback_text and _text_quality_score(fallback_text) > _text_quality_score(text):
                text = fallback_text
        return clean_extracted_text(text)

    if suffix == ".docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Install python-docx to process DOCX uploads.",
            ) from exc

        document = Document(str(path))
        return clean_extracted_text("\n".join(paragraph.text for paragraph in document.paragraphs))

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"Unsupported document type: {suffix}",
    )


def _read_spreadsheet(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Install openpyxl to process XLSX uploads.",
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append(" | ".join(values))
    finally:
        workbook.close()

    return clean_extracted_text("\n".join(lines))


def _read_image(path: Path) -> str:
    try:
        from PIL import Image
        import pytesseract
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Install pillow and pytesseract to process image OCR uploads.",
        ) from exc

    try:
        with Image.open(path) as image:
            return clean_extracted_text(pytesseract.image_to_string(image))
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Unable to extract text from image upload: {exc}",
        ) from exc


def _download_headers(
    headers: dict[str, str] | None,
    api_key: str | None,
    bearer_token: str | None,
) -> dict[str, str]:
    resolved_headers = {"User-Agent": "Teamoria-AI-Service/1.0"}

    effective_api_key = api_key or settings.backend_file_api_key
    if effective_api_key:
        resolved_headers[settings.backend_file_api_key_header or "x-api-key"] = effective_api_key

    effective_bearer_token = bearer_token or settings.backend_file_bearer_token
    if effective_bearer_token:
        resolved_headers["Authorization"] = f"Bearer {effective_bearer_token}"

    resolved_headers.update(headers or {})
    return resolved_headers


def clean_extracted_text(text: str) -> str:
    """Normalize extracted/transcribed text before AI analysis and persistence."""

    return _clean_extracted_text(text)


def _read_pdf_with_pypdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Install pypdf to process PDF uploads.",
        ) from exc

    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages).strip()


def _read_pdf_with_pymupdf(path: Path) -> str:
    try:
        import fitz
    except ImportError:
        return ""

    with fitz.open(str(path)) as document:
        return "\n".join(page.get_text("text") or "" for page in document).strip()


def _clean_extracted_text(text: str) -> str:
    candidate = _maybe_fix_mojibake(text)
    candidate = unicodedata.normalize("NFKC", candidate)
    candidate = candidate.replace("\u200f", " ").replace("\u200e", " ")
    candidate = _collapse_spaced_latin_glyphs(candidate)
    candidate = re.sub(r"[ \t]+", " ", candidate)
    candidate = re.sub(r"\n{3,}", "\n\n", candidate)
    return candidate.strip()


def _collapse_spaced_latin_glyphs(text: str) -> str:
    return "\n".join(_collapse_spaced_latin_line(line) for line in text.splitlines())


def _collapse_spaced_latin_line(line: str) -> str:
    tokens = line.split()
    if not _looks_like_spaced_latin_glyphs(tokens):
        return line

    collapsed = []
    previous = ""

    for token in tokens:
        current = token
        if _needs_space_between_spaced_glyphs(previous, current):
            collapsed.append(" ")
        collapsed.append(current)
        previous = current

    line = "".join(collapsed)
    line = re.sub(r"\s+([,.;:!?%)])", r"\1", line)
    line = re.sub(r"([(])\s+", r"\1", line)
    line = re.sub(r"\s*-\s*", "-", line)
    line = re.sub(r"([a-z])([A-Z])", r"\1 \2", line)
    line = re.sub(r"([A-Za-z])(\d)", r"\1 \2", line)
    line = re.sub(r"(\d)([A-Z])", r"\1 \2", line)
    line = re.sub(r"\b([A-Z]{2,})([A-Z][a-z])", r"\1 \2", line)
    line = re.sub(r"\b([A-Za-z]{4,})(in|of|to|for|and|with)(?=[A-Z\s]|$)", r"\1 \2", line)
    return line


def _looks_like_spaced_latin_glyphs(tokens: list[str]) -> bool:
    if len(tokens) < 8:
        return False

    glyph_tokens = [
        token
        for token in tokens
        if len(token) == 1 and re.match(r"[A-Za-z0-9'&+/#@._-]", token)
    ]
    latin_glyph_tokens = [
        token
        for token in glyph_tokens
        if re.match(r"[A-Za-z0-9]", token)
    ]
    return len(glyph_tokens) / len(tokens) >= 0.7 and len(latin_glyph_tokens) >= 6


def _needs_space_between_spaced_glyphs(previous: str, current: str) -> bool:
    if not previous:
        return False
    if current in ".,;:!?%)":
        return False
    if previous in "([/@._-":
        return False
    if current in "([/@._-":
        return True
    if previous.islower() and current.isupper():
        return True
    if previous.isalpha() and current.isdigit():
        return True
    if previous.isdigit() and current.isupper():
        return True
    return False


def _maybe_fix_mojibake(text: str) -> str:
    if not text:
        return ""

    try:
        candidate = text.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
    except UnicodeError:
        return text

    if _arabic_char_count(candidate) > _arabic_char_count(text) and len(candidate) > len(text) * 0.35:
        return candidate

    return text


def _is_low_quality_text(text: str) -> bool:
    clean_text = text.strip()
    if len(clean_text) < 40:
        return True
    if _has_spaced_latin_glyph_artifacts(clean_text):
        return True
    return _text_quality_score(clean_text) < 0


def _has_spaced_latin_glyph_artifacts(text: str) -> bool:
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return False

    artifact_lines = sum(
        1
        for line in lines
        if _looks_like_spaced_latin_glyphs(line.split())
    )
    return artifact_lines / len(lines) >= 0.25


def _text_quality_score(text: str) -> int:
    return _arabic_char_count(_clean_without_mojibake_retry(text)) - _mojibake_marker_count(text) * 3


def _clean_without_mojibake_retry(text: str) -> str:
    return unicodedata.normalize("NFKC", text or "")


def _arabic_char_count(text: str) -> int:
    return sum(1 for char in text if "\u0600" <= char <= "\u06ff")


def _mojibake_marker_count(text: str) -> int:
    return sum(text.count(marker) for marker in ("Ø", "Ù", "ï", "�"))
